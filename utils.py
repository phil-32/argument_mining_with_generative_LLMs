# -*- coding: utf-8 -*-
from tqdm import tqdm
import logging
import fuzzysearch
import pandas as pd
import difflib
import os
import openpyxl
import traceback
from openpyxl.styles import PatternFill, Alignment
import re
import json
from argument_mining_persuade import prompt_generation

log = logging.getLogger('main_logger')


def invoke_llm(df_essay, df_du, prompt_var_func_dict, parser_func,
               fuzzysearch_factor, prompt_module, model, prompt_format,
               example_df_file, example_format_function,
               output_dir="", batch_result_file=""):
    """
    - Setup logging
    - Build prompts
    - Invoke the LLM or read a batch result
    - Start functions to calculate the metrics
    """
    stats_dict = {}
    stats_dict["total_classified_du"] = 0
    stats_dict["total_verbatim_du"] = 0
    stats_dict["total_matched_not_verbatim_du"] = 0
    stats_dict["syntax_err_list"] = []
    result_df = pd.DataFrame(
        columns=["essay_id", "discourse_start", "discourse_end",
                 "discourse_text", "discourse_type", "original_essay_text",
                 "llm_output"])
    prompt_log = createLogger(output_dir, "prompts")

    for i, df_essay_row in tqdm(df_essay.iterrows(), total=df_essay.shape[0]):
        log.debug("")
        log.debug(f"## Processing_essay {df_essay_row.essay_id} ##")

        prompt_dict = prompt_module.prompt_dict
        if prompt_format == "simple":
            prompt = prompt_generation.assembleSimplePrompt(
                example_df_file=example_df_file,
                example_format_function=example_format_function,
                essay_text=df_essay_row.full_text_clean,
                **prompt_dict)
        elif prompt_format == "chat":
            prompt = prompt_generation.assembleChatPrompt(
                example_df_file=example_df_file,
                example_format_function=example_format_function,
                essay_text=df_essay_row.full_text_clean,
                **prompt_dict)
        elif prompt_format == "CoT":
            prompt = prompt_generation.assembleCoTPrompt(
                example_df_file=example_df_file,
                example_format_function=example_format_function,
                essay_text=df_essay_row.full_text_clean,
                **prompt_dict)

        chain = prompt | model

        # Get additonal logger for prompts
        prompt_log.debug(f"## Processing_essay {df_essay_row.essay_id} ##")

        # Set current essay text
        prompt_var_dict = fill_prompt_var_dict(df_essay_row,
                                               prompt_var_func_dict)
        # Write filled in prompt to log
        filled_prompt = remove_special_characters(
            chain.first.format(**prompt_var_dict))
        prompt_log.debug(
            f"--Full prompt:\n{filled_prompt}")
        prompt_log.debug("--End of prompt\n")

        # Directly invoke chain
        if not batch_result_file:
            # Invoke chain and get output string
            message = chain.invoke(prompt_var_dict)
            output = message.content
            # Write token usage to log
            prompt_log.debug(message.response_metadata)
        else:
            output = llm_output_from_batch_file(
                batch_result_file, df_essay_row.essay_id)

        # print Output also to prompt log
        prompt_log.debug(f"The LLM output was: "
                         f"\n\n{remove_special_characters(output)}\n\n")

        stats_dict, result_df = parse_llm_result(
            stats_dict=stats_dict,
            essay_text=df_essay_row.full_text_clean,
            essay_id=df_essay_row.essay_id,
            parser_func=parser_func,
            llm_output=output,
            result_df=result_df,
            fuzzysearch_factor=fuzzysearch_factor)

    statistics_df = pd.DataFrame({
        "Total DUs in ground truth":
            len(df_du[df_du["essay_id"].isin(df_essay.essay_id)]),
        "Total classified DUs": stats_dict["total_classified_du"],
        "Total usable discourse units":
            stats_dict["total_verbatim_du"]
            + stats_dict["total_matched_not_verbatim_du"],
        "verbatim DUs": stats_dict["total_verbatim_du"],
        "Total non-verbatim DUs":
            stats_dict["total_classified_du"]
            - stats_dict["total_verbatim_du"],
        "Non-verbatim DUs matched with fuzzysearch":
            stats_dict["total_matched_not_verbatim_du"],
        "Total Number of essays": len(df_essay),
        "Number of unparsable essays": len(stats_dict["syntax_err_list"]),
        }, index=[0])

    if output_dir:
        statistics_df.to_csv(os.path.join(output_dir, "statistics_df.csv"))
        result_df.to_csv(os.path.join(output_dir, "result_df.csv"))

    log.info(f"\nParsing info:\n"
             f"{statistics_df.transpose().to_string(header=False)}\n")

    return result_df, statistics_df


def parse_llm_result(stats_dict, essay_text, essay_id, parser_func, llm_output,
                     result_df, fuzzysearch_factor):
    """
    Parse the output of the LLM with the parser_func and create a result
    dataframe.
    """
    try:
        # Use parser to get output dict like {<du_text>: <du_type>}
        output_dict = parser_func(llm_output)

        verbatim_counter = 0
        matched_not_verbatim_counter = 0
        for span, du_type in output_dict.items():
            # Write discourse units, that were generated verbatim by LLM
            # and write them to new result df
            if span in essay_text:
                verbatim_counter += 1
                result_df = pd.concat(
                    [result_df, pd.DataFrame(
                        {"essay_id": essay_id,
                         "discourse_start": essay_text.find(span),
                         "discourse_end":
                             essay_text.find(span) + len(span),
                         "discourse_text": span,
                         "discourse_type": parse_discourse_type(du_type),
                         "llm_output": llm_output,
                         "original_essay_text": essay_text,
                         }, index=[0])],
                    ignore_index=True)
            # Use fuzzysearch to match non-verbatim discourse units and
            # also write them to result df
            else:
                log.debug(f"\nmatch_not_verbatim_du() of span:\n"
                          f"{remove_special_characters(span)}")
                matches = match_not_verbatim_du(
                    span, essay_text, fuzzysearch_factor)
                if len(matches) == 1:
                    matched_not_verbatim_counter += 1
                    log.debug("+ match of non-verbatim span successful")
                    result_df = pd.concat(
                        [result_df, pd.DataFrame(
                            {"essay_id": essay_id,
                             "discourse_start": matches[0].start,
                             "discourse_end": matches[0].end,
                             "discourse_text": matches[0].matched,
                             "discourse_type":
                                 parse_discourse_type(du_type),
                             "llm_output": llm_output,
                             "original_essay_text": essay_text,
                             "llm_prompt": essay_text,
                             }, index=[0])],
                        ignore_index=True)
                else:
                    log.debug("- non-verbatim span could not be matched")
        log.debug(f"\nVerbatim discourse units in essay: "
                  f"{verbatim_counter} of {len(output_dict)}")
        log.debug(f"Additionally matched non verbatim discourse units: "
                  f"{matched_not_verbatim_counter} "
                  f"of {len(output_dict) - verbatim_counter}")

        stats_dict["total_verbatim_du"] += verbatim_counter
        stats_dict["total_matched_not_verbatim_du"] += \
            matched_not_verbatim_counter
        stats_dict["total_classified_du"] += len(output_dict)
        if stats_dict["total_classified_du"] == 0:
            log.debug(f"No discourse units were classified for the essay. "
                      f"The LLM output was: \n\n{llm_output}\n\n")
    # Count non parsable LLM output
    except SyntaxError:
        log.debug(f"SyntaxError in LLM output! LLM output is:"
                  f"\n\n{remove_special_characters(llm_output)}\n\n")
        stats_dict["syntax_err_list"].append(llm_output)
    except Exception as e:
        log.debug(f"The follogwing error while parsing the LLM output:\n\n"
                  f"{e}\n\n{traceback.format_exc()}\n\nThe LLM output "
                  f"was:\n\n{remove_special_characters(llm_output)}\n\n")
        stats_dict["syntax_err_list"].append(llm_output)

    return stats_dict, result_df


def remove_special_characters(span):
    """
    Remove any special characters which might cause problems with logging.
    """
    # remove any special non ascii characters
    strencode = span.encode("ascii", "ignore")
    return strencode.decode()


def write_batch_file(df_essay, df_du, prompt_var_func_dict, parser_func,
                     prompt_module, model, prompt_format,
                     example_df_file, example_format_function, exp_string="",
                     output_dir="", numOfEssaysPerBatchFile=None):
    """
    Write batch file for OpenAI Batch-API
    """
    tasks = []

    for i, df_essay_row in tqdm(df_essay.iterrows(),
                                total=df_essay.shape[0]):
        prompt_var_dict = fill_prompt_var_dict(df_essay_row,
                                               prompt_var_func_dict)
        # prompt = chain.first.format(**prompt_var_dict)

        prompt_dict = prompt_module.prompt_dict
        if prompt_format == "simple":
            prompt = prompt_generation.assembleSimplePrompt(
                example_df_file=example_df_file,
                example_format_function=example_format_function,
                essay_text=df_essay_row.full_text_clean,
                **prompt_dict)
        elif prompt_format == "chat":
            prompt = prompt_generation.assembleChatPrompt(
                example_df_file=example_df_file,
                example_format_function=example_format_function,
                essay_text=df_essay_row.full_text_clean,
                **prompt_dict)
        elif prompt_format == "CoT":
            prompt = prompt_generation.assembleCoTPrompt(
                example_df_file=example_df_file,
                example_format_function=example_format_function,
                essay_text=df_essay_row.full_text_clean,
                **prompt_dict)

        chain = prompt | model

        # Create message list in correct format for json dump
        messages = []
        if chain.first.__class__.__name__ == 'ChatPromptTemplate':
            for message in chain.first.format_messages(**prompt_var_dict):
                # Translate keys used by langchain to the roles used by OpenAI
                if message.type == "system":
                    role = "system"
                elif message.type == "ai":
                    role = "assistant"
                elif message.type == "human":
                    role = "user"
                else:
                    raise Exception("Unknown or not inplemented message type.")
                messages.append({
                    "role": role,
                    "content": message.content})
        elif chain.first.__class__.__name__ == 'PromptTemplate':
            messages.append({
                "role": "user",
                "content": chain.first.format(**prompt_var_dict)})

        task = {
            "custom_id": df_essay_row.essay_id,
            "method": "POST",
            "url": "/v1/chat/completions",
            "body": {
                "model": chain.last.model_name,
                "temperature": chain.last.temperature,
                "max_tokens": chain.last.max_tokens,
                "messages": messages,
                }
            }
        tasks.append(task)

    # Creating the file
    if numOfEssaysPerBatchFile is None:
        file_name = os.path.join(output_dir, f"batch_{exp_string}.jsonl")
        with open(file_name, 'w') as file:
            for obj in tasks:
                file.write(json.dumps(obj) + '\n')
    elif numOfEssaysPerBatchFile > 0:
        fileNr = 0
        for i, obj in enumerate(tasks):
            file_name = os.path.join(output_dir,
                                     f"batch_{exp_string}_{fileNr}.jsonl")
            if (i + 1) % numOfEssaysPerBatchFile == 0:
                fileNr += 1
            with open(file_name, 'a') as file:
                file.write(json.dumps(obj) + '\n')
    else:
        raise Exception("numOfEssaysPerBatchFile must be None or positve "
                        "integer!")


def llm_output_from_batch_file(batch_result_file, essay_id):
    """
    Retrieve output generated by the OpenAI model for a specified essay
    from a batch file.

    Parameters
    ----------
    batch_result_file : str
        Path to batch file.
    essay_id : str
        essay_id of the essay which was classified.

    Returns
    -------
    str
        llm_output for the essay, which was to be classified.
    """
    with open(batch_result_file, 'r') as file:
        for line in file:
            # Parsing the JSON string into a dict and appending to the list of
            # results
            json_object = json.loads(line.strip())
            if json_object['custom_id'] == essay_id:
                return json_object[
                    'response']['body']['choices'][0]['message']['content']


def write_metrics_statistics_to_excel_overview(
        statistics_df, metrics_df, output_dir, overview_file_name, exp_string):
    """
    Write metrics of the current experiment to an overview excel file.
    """
    if metrics_df is None:
        return
    # Prepare data frame row by flattening and concatenating dfs
    columns = [f"{du_type} (span)" for du_type in metrics_df.columns]
    columns.extend([f"{du_type} (word)" for du_type in metrics_df.columns])
    columns.extend(statistics_df.columns)

    row_values_df = pd.concat([metrics_df.loc["F1_span"],
                               metrics_df.loc["F1_word"],
                               statistics_df.transpose()], ignore_index=True)
    output_df = pd.DataFrame(row_values_df.values.transpose(), columns=columns,
                             index=[exp_string])
    output_df = output_df.transpose()
    file_path = os.path.join(output_dir, overview_file_name)
    if os.path.isfile(file_path):
        read_df = pd.read_excel(file_path, header=0, index_col=0)
        output_df = pd.concat([read_df, output_df], axis=1)
        output_df.to_excel(file_path, sheet_name="overview",
                           float_format="%.2f")
    else:
        output_df.to_excel(file_path, sheet_name="overview",
                           float_format="%.2f")

    # Basic formatting of excel sheet with openpyxl
    wb = openpyxl.load_workbook(file_path)
    ws = wb.worksheets[0]
    for row in (9, 17):
        for cell in ws[row]:
            cell.fill = PatternFill("solid", fgColor="FFFF00")
    for cell in ws[1]:
        cell.alignment = Alignment(text_rotation=30)
    for row in range(1, ws.max_row + 1):
        cell = ws[f'A{row}']
        cell.alignment = Alignment(horizontal="left")
    ws.column_dimensions["A"].width = 35
    wb.save(file_path)


def parser_python_dict(llm_output):
    llm_output = llm_output.replace(
        r"```python", "").replace("```", "").strip()
    return eval(llm_output)


def parser_XML(llm_output):
    output_dict = {}
    matches = re.finditer(r"<\/?([^>]+)>", llm_output)
    last_tag = ""
    last_end_pos = ""
    for match in matches:
        tag = match.group(1)
        start_pos = match.start()
        end_pos = match.end()
        # Check if it a start or an end tag.
        if len(re.findall(r"<\/([^>]+)>", match.group())) == 0:
            is_start_tag = True
        # If it is an end tag check if it is similar to the last start tag
        else:
            is_start_tag = False
            if not is_start_tag and tag == last_tag:
                output_dict[llm_output[last_end_pos: start_pos]] = \
                    parse_discourse_type(tag)

        # Set state for next iteration
        last_tag = tag
        last_end_pos = end_pos
    return output_dict


def parser_TANL(llm_output):
    output_dict = {}
    matches = re.finditer(r"([^\[\]\|]+)\|([^\]]+)\]", llm_output)
    for match in matches:
        output_dict[match.group(1)] = match.group(2)
    return output_dict


def match_not_verbatim_du(span, full_essay, fuzzysearch_factor=10):
    """
    Invoke fuzzysearch matching
    """
    matches = fuzzysearch.find_near_matches(
        span, full_essay, max_l_dist=int(len(span) / fuzzysearch_factor))
    return matches


def fill_prompt_var_dict(df_row, prompt_var_func_dict):
    """
    Create a Python dictionary where the keys are the possible prompt variables
    and the values are retrieved from the df_row using a function defined in
    prompt_var_func_dict which specifies how to retrieve the value from the
    dataframe.
    """
    d = {}
    for var, func in prompt_var_func_dict.items():
        d[var] = func(df_row)
    return d


def get_valid_types_list():
    return ["Lead", "Position", "Claim", "Counterclaim", "Rebuttal",
            "Evidence", "Concluding Statement"]


def parse_discourse_type(classified_type):
    """
    Parse the discourse unit types specified by the LLM, e.g., in the
    XML-like syntax. This allows slight variations in the spelling or
    variations in the capitalization of the discourse unit types.
    """
    valid_types = get_valid_types_list()
    if classified_type:
        matches = difflib.get_close_matches(classified_type.lower(),
                                            valid_types, n=1, cutoff=0.6)
        if len(matches) == 1:
            return matches[0]
    return None


def add_predictionstring_span(df, full_text_col, du_start_index_col,
                              du_end_index_col):
    """
    Build the prediciton string like it is used in the Kaggle competition
    """
    prediction_df = df.copy()
    for i, row in prediction_df.iterrows():
        text = row[full_text_col]
        start_idx = row[du_start_index_col]
        end_idx = row[du_end_index_col]
        prediction_df.at[i, "predictionstring"] = _get_start_and_end_word_idx(
            text, start_idx, end_idx)
    return prediction_df


def build_predictionstring_df_word(df, full_text_col, du_start_index_col,
                                   du_end_index_col, essay_id_col,
                                   discourse_type_col):
    word_predictionstring_df = pd.DataFrame(
        columns=["essay_id", "discourse_type", "predictionstring"])
    for i, row in df.iterrows():
        text = row[full_text_col]
        start_idx = row[du_start_index_col]
        end_idx = row[du_end_index_col]
        index_string = _get_start_and_end_word_idx(text, start_idx, end_idx)
        for idx in index_string.split():
            word_predictionstring_df = pd.concat(
                [word_predictionstring_df, pd.DataFrame(
                    {"essay_id": row.essay_id,
                     "discourse_type": row.discourse_type,
                     "predictionstring": idx
                     }, index=[0])],
                ignore_index=True)
    return word_predictionstring_df


def _get_start_and_end_word_idx(text, start_idx, end_idx):
    word_start_idx = len(text[:start_idx].split())
    word_end_idx = len(text[:end_idx].split())
    return " ".join(map(str, range(word_start_idx, word_end_idx)))


def check_overlap(s1, e1, s2, e2):
    """
    Calculate the overlap between 2 sequences defined by their start and
    end indices

    Parameters
    ----------
    s1 : int
        Start index of 1st sequence
    e1 : int
        End index of 1st sequence
    s2 : int
        Start index of 2nd sequence
    e2 : int
        End index of 2nd sequence

    Returns
    -------
    None if there is no overlap
    Tuple of integers with the start and end index of the overlap if there
    is any
    """
    # Always take start and end of the longer segment as s1 and e1
    if e1 - s1 < s2 - e2:
        temp_s1 = s1
        temp_e1 = e1
        s1 = s2
        e1 = e2
        s2 = temp_s1
        e2 = temp_e1
    # is start of sequence 2 within sequence 1?
    if s1 <= s2 and s2 < e1:
        # is end of sequence 2 also within sequence 1?
        if e2 <= e1:
            return (s2, e2)
        else:
            return (s2, e1)
    else:
        return None


def createLogger(full_output_dir, logger_name="main_logger"):
    """
    Setup logging (to console and to file)
    """
    log = logging.getLogger(logger_name)
    log.setLevel(logging.DEBUG)
    formatter = logging.Formatter('%(message)s')
    log.propagate = False
    # Create a console handler for logging INFO level messages
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(formatter)
    # Create a file handler for logging DEBUG level messages
    if not os.path.exists(full_output_dir):
        os.makedirs(full_output_dir)
    log_file = os.path.join(full_output_dir, f"{logger_name}.log")
    file_handler = logging.FileHandler(log_file, mode="w")
    # clear old loggers
    if log.hasHandlers():
        log.handlers.clear()
    # Add handlers to the logger
    log.addHandler(console_handler)
    log.addHandler(file_handler)
    return log
